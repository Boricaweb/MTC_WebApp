import openpyxl
import json
import sys
import io
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, 'อัปเดตตารางงานซ่อมติดตามรายเดือน.xlsm')

wb = openpyxl.load_workbook(excel_path, data_only=True)

# Focus on the key sheets
key_sheets = [
    '🛠️งานซ่อมตามแผน',
    '🔧รายการงานซ่อมทั้งหมด',
    '📈กราฟรวมงานซ่อม',
    '📊วิเคราะห์งานรายเดือน',
    '💻Update งานซ่อม (เมษายน)',
    '📆สรุปรายปี2569',
    '📦Stock',
    'คลังงานซ่อม',
]

for name in key_sheets:
    ws = wb[name]
    print(f'\n=== Sheet: {name} ({ws.max_row}r x {ws.max_column}c) ===')
    max_rows = min(8, ws.max_row)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=min(20, ws.max_column), values_only=False), 1):
        vals = [str(cell.value)[:80] if cell.value is not None else '' for cell in row]
        if any(v for v in vals):
            print(f'  Row {row_idx}: {json.dumps(vals, ensure_ascii=False)}')
